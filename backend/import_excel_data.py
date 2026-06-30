"""Import Excel data into PostgreSQL database"""
import sys
import os
from pathlib import Path
import openpyxl
from datetime import datetime

# Add the backend directory to path
sys.path.insert(0, str(Path(__file__).parent))

from sqlalchemy.orm import Session
from app.core.database import get_db, engine, Base
from app.models.resource import Resource, ResourceStatus, Allocation
from app.models.project import Project
from app.models.asset_type import AssetType
from app.models.user import User
import uuid


def create_or_get_asset_type(session: Session, name: str, description: str) -> AssetType:
    """Create or get existing asset type"""
    asset_type = session.query(AssetType).filter(AssetType.name == name).first()
    if not asset_type:
        asset_type = AssetType(
            name=name,
            description=description,
            is_active=True
        )
        session.add(asset_type)
        session.commit()
        session.refresh(asset_type)
        print(f"Created asset type: {name}")
    return asset_type


def get_or_create_default_project(session: Session, name: str = "General Infrastructure", description: str = "Default project for imported infrastructure resources") -> Project:
    """Get or create default project"""
    project = session.query(Project).filter(Project.name == name).first()
    if not project:
        # Get admin user as owner
        admin_user = session.query(User).filter(User.username == "admin").first()
        if not admin_user:
            print("Error: Admin user not found. Please run seed_db.py first.")
            return None
        
        project = Project(
            project_code="INF-001",
            name=name,
            description=description,
            status="Active",
            budget=1000000.00,
            allocated_budget=0.00,
            owner_id=admin_user.id,
            created_by=admin_user.id
        )
        session.add(project)
        session.commit()
        session.refresh(project)
        print(f"Created project: {name}")
    return project


def import_servers_from_excel(session: Session, excel_path: str):
    """Import servers from Excel file"""
    print(f"\n=== Importing servers from {excel_path} ===")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['servers']
    
    # Get headers
    headers = [str(cell.value).strip().lower().replace(' ', '_').replace('(', '').replace(')', '') for cell in ws[1]]
    print(f"Headers: {headers}")
    
    asset_type = create_or_get_asset_type(session, "Server", "Server/Compute resources")
    project = get_or_create_default_project(session)
    
    # Get admin user
    admin_user = session.query(User).filter(User.username == "admin").first()
    if not admin_user:
        print("Error: Admin user not found. Please run seed_db.py first.")
        return
    
    imported_count = 0
    skipped_count = 0
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        row_data = {}
        for i, cell in enumerate(row):
            if i < len(headers):
                row_data[headers[i]] = cell.value
        
        # Extract required fields
        serial_number = row_data.get('machine_serial_number', '')
        make = row_data.get('make', '')
        model = row_data.get('model', '')
        processor = row_data.get('processor_type', '')
        ram = row_data.get('ram_size_gb', '')
        storage = row_data.get('primary_storage__type_&_size_tb', '')
        os = row_data.get('os_with_version', '')
        warranty_end = row_data.get('warrenty_end_date', None)
        
        if not serial_number or not make:
            print(f"Row {row_idx}: Missing serial number or make, skipping")
            skipped_count += 1
            continue
        
        # Calculate cost based on row index as placeholder
        cost = 50000.00 + (row_idx * 10000)  # Placeholder cost
        
        # Handle warranty date
        warranty_date_str = ""
        if warranty_end:
            if hasattr(warranty_end, 'strftime'):
                warranty_date_str = warranty_end.strftime('%Y-%m-%d')
            else:
                warranty_date_str = str(warranty_end)
        
        # Create custom field values
        custom_fields = {
            "serial_number": str(serial_number),
            "brand": str(make),
            "model": str(model),
            "processor": str(processor) if processor else "",
            "ram": str(ram) if ram else "",
            "storage": str(storage) if storage else "",
            "os": str(os) if os else "",
            "warranty_end": warranty_date_str,
            "room_number": row_data.get('present__room_no', ''),
            "system_role": row_data.get('system_role', ''),
            "specific_task": row_data.get('specific_task,_if_any', ''),
        }
        
        resource = Resource(
            project_id=project.id,
            asset_type_id=asset_type.id,
            name=f"{make} {model} - {serial_number}",
            cost=cost,
            allocation_date=datetime.now().date(),
            status=ResourceStatus.ACTIVE,
            custom_field_values=custom_fields,
            created_by=admin_user.id,
            created_at=datetime.utcnow(),
            updated_at=datetime.utcnow()
        )
        
        session.add(resource)
        session.flush()
        
        # Create allocation
        allocation = Allocation(
            resource_id=resource.id,
            project_id=project.id,
            cost_at_allocation=cost,
            created_by=admin_user.id
        )
        session.add(allocation)
        
        imported_count += 1
        print(f"  Row {row_idx}: Imported {make} {model}")
    
    session.commit()
    print(f"\nImported {imported_count} servers, skipped {skipped_count}")


def import_seaw_overview(session: Session, excel_path: str):
    """Import seaw compute overview data"""
    print(f"\n=== Importing seaw overview from {excel_path} ===")
    
    wb = openpyxl.load_workbook(excel_path)
    ws = wb['seaw']
    
    # Parse the overview data
    overview_data = {}
    for row in ws.iter_rows(min_row=2):
        key = row[0].value
        value = row[1].value
        if key:
            overview_data[key] = value
    
    print(f"Overview data: {overview_data}")
    
    # Create a summary resource for seaw compute
    asset_type = create_or_get_asset_type(session, "Seaw Compute", "Seaw compute cluster")
    project = get_or_create_default_project(session)
    
    admin_user = session.query(User).filter(User.username == "admin").first()
    if not admin_user:
        print("Error: Admin user not found")
        return
    
    # Create a summary resource
    summary = overview_data.get('No. of servers', '0')
    gpu_info = overview_data.get('No. of GPU', '0')
    
    resource = Resource(
        project_id=project.id,
        asset_type_id=asset_type.id,
        name=f"Seaw Compute Summary - {summary}",
        cost=500000.00,
        allocation_date=datetime.now().date(),
        status=ResourceStatus.ACTIVE,
        custom_field_values={
            "summary": summary,
            "gpu_count": gpu_info,
            "imported_from": "data.xlsx",
            "import_date": datetime.now().strftime('%Y-%m-%d'),
        },
        created_by=admin_user.id,
        created_at=datetime.utcnow(),
        updated_at=datetime.utcnow()
    )
    
    session.add(resource)
    session.flush()
    
    allocation = Allocation(
        resource_id=resource.id,
        project_id=project.id,
        cost_at_allocation=500000.00,
        created_by=admin_user.id
    )
    session.add(allocation)
    
    session.commit()
    print(f"Created seaw compute summary: {summary}")


def main():
    """Main function to import all Excel data"""
    print("=" * 60)
    print("Excel Data Import Script")
    print("=" * 60)
    
    excel_folder = "d:/Anugya_PRMS/data/excel_folder"
    data_xlsx = os.path.join(excel_folder, "data.xlsx")
    
    if not os.path.exists(data_xlsx):
        print(f"Error: Excel file not found at {data_xlsx}")
        return
    
    # Create all tables
    Base.metadata.create_all(bind=engine)
    print("Database tables created/verified")
    
    # Get database session
    session = next(get_db())
    
    try:
        # Import seaw overview
        import_seaw_overview(session, data_xlsx)
        
        # Import servers
        import_servers_from_excel(session, data_xlsx)
        
        print("\n" + "=" * 60)
        print("Import completed successfully!")
        print("=" * 60)
        
    except Exception as e:
        session.rollback()
        print(f"Error during import: {e}")
        import traceback
        traceback.print_exc()
    finally:
        session.close()


if __name__ == "__main__":
    main()
