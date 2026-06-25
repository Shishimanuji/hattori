#!/usr/bin/env python
"""Generate sample Excel file for PRMS dashboard"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Create a new workbook
wb = Workbook()

# Remove default sheet
wb.remove(wb.active)

# ==================== RESOURCES SHEET ====================
ws_resources = wb.create_sheet("Resources")

# Define headers for Resources sheet
resource_headers = ["Name", "Cost", "AssetType", "Status", "Project"]
ws_resources.append(resource_headers)

# Sample resource data
resource_data = [
    # IT Infrastructure Resources
    ("Dell Server DS-2024-01", 45000.00, "Equipment", "Active", "Enterprise Infrastructure"),
    ("Cisco Network Switch CS-8K", 35000.00, "Equipment", "Active", "Enterprise Infrastructure"),
    ("UPS Backup System UPS-500KVA", 28000.00, "Equipment", "Active", "Enterprise Infrastructure"),
    ("Server Rack Mount SR-42U", 12000.00, "Equipment", "Active", "Enterprise Infrastructure"),
    ("Network Cables & Fiber", 8500.00, "Equipment", "Active", "Enterprise Infrastructure"),
    
    # Software Licenses
    ("Microsoft Office 365 Licenses", 15000.00, "Software", "Active", "Digital Transformation"),
    ("Adobe Creative Suite Annual", 22000.00, "Software", "Active", "Digital Transformation"),
    ("Database Enterprise Edition", 50000.00, "Software", "Active", "Data Analytics Platform"),
    ("Cloud Infrastructure Credits", 75000.00, "Software", "Active", "Cloud Migration"),
    ("Security & Monitoring Tools", 35000.00, "Software", "Active", "Cybersecurity Initiative"),
    
    # Personnel - Full-time
    ("Senior Architect - John Smith", 180000.00, "Personnel", "Active", "Cloud Migration"),
    ("Project Manager - Sarah Wilson", 120000.00, "Personnel", "Active", "Digital Transformation"),
    ("DevOps Engineer - Mike Chen", 140000.00, "Personnel", "Active", "Enterprise Infrastructure"),
    ("Data Scientist - Emily Rodriguez", 150000.00, "Personnel", "Active", "Data Analytics Platform"),
    ("Security Lead - David Kumar", 165000.00, "Personnel", "Active", "Cybersecurity Initiative"),
    ("Frontend Developer - Lisa Anderson", 110000.00, "Personnel", "Active", "Digital Transformation"),
    ("Backend Developer - James Brown", 130000.00, "Personnel", "Active", "Enterprise Infrastructure"),
    
    # Additional Equipment
    ("Backup Storage Array", 55000.00, "Equipment", "Active", "Data Analytics Platform"),
    ("Workstations for Dev Team", 32000.00, "Equipment", "Active", "Enterprise Infrastructure"),
    ("Video Conference System", 18000.00, "Equipment", "Inactive", "Digital Transformation"),
]

# Style the header row
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for cell in ws_resources[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Add resource data
for row in resource_data:
    ws_resources.append(row)

# Format data rows
for row in ws_resources.iter_rows(min_row=2, max_row=len(resource_data) + 1):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # Format cost column as currency
        if cell.column == 2:  # Cost column
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")

# Adjust column widths
ws_resources.column_dimensions['A'].width = 35
ws_resources.column_dimensions['B'].width = 15
ws_resources.column_dimensions['C'].width = 15
ws_resources.column_dimensions['D'].width = 12
ws_resources.column_dimensions['E'].width = 30

# ==================== PROJECTS SHEET ====================
ws_projects = wb.create_sheet("Projects")

# Define headers for Projects sheet
project_headers = ["Name", "Budget", "Status", "StartDate", "Description"]
ws_projects.append(project_headers)

# Sample project data with specific budget allocations to meet requirements
base_date = datetime.now()

project_data = [
    # Active Project - 85% utilization (high alert)
    ("Enterprise Infrastructure", 250000.00, "Active", 
     (base_date - timedelta(days=180)).strftime("%Y-%m-%d"),
     "Upgrade core IT infrastructure with new servers and networking equipment"),
    
    # Active Project - 100% utilization (critical alert)
    ("Cloud Migration", 300000.00, "Active",
     (base_date - timedelta(days=120)).strftime("%Y-%m-%d"),
     "Migrate on-premises systems to cloud infrastructure"),
    
    # Active Project - 60% utilization (normal)
    ("Digital Transformation", 200000.00, "Active",
     (base_date - timedelta(days=150)).strftime("%Y-%m-%d"),
     "Implement modern applications and digital tools"),
    
    # Active Project - 70% utilization (normal)
    ("Data Analytics Platform", 280000.00, "Active",
     (base_date - timedelta(days=90)).strftime("%Y-%m-%d"),
     "Build comprehensive data analytics and reporting platform"),
    
    # Pending Project
    ("AI & Machine Learning Initiative", 400000.00, "Pending",
     (base_date + timedelta(days=30)).strftime("%Y-%m-%d"),
     "Develop AI-powered solutions for business optimization"),
    
    # Pending Project
    ("Mobile Application Dev", 150000.00, "Pending",
     (base_date + timedelta(days=45)).strftime("%Y-%m-%d"),
     "Create mobile apps for iOS and Android platforms"),
    
    # On Hold Project
    ("Legacy System Modernization", 350000.00, "On Hold",
     (base_date - timedelta(days=60)).strftime("%Y-%m-%d"),
     "Refactor legacy monolithic applications to microservices"),
    
    # Completed Project
    ("Cybersecurity Initiative", 200000.00, "Completed",
     (base_date - timedelta(days=300)).strftime("%Y-%m-%d"),
     "Comprehensive security audit and implementation of security controls"),
    
    # Active Project - 50% utilization (normal)
    ("Business Automation", 180000.00, "Active",
     (base_date - timedelta(days=75)).strftime("%Y-%m-%d"),
     "Automate routine business processes and workflows"),
    
    # Pending Project
    ("Customer Portal Redesign", 220000.00, "Pending",
     (base_date + timedelta(days=60)).strftime("%Y-%m-%d"),
     "Redesign customer-facing portal with modern UX/UI"),
]

# Style the header row
for cell in ws_projects[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

# Add project data
for row in project_data:
    ws_projects.append(row)

# Format data rows
for row in ws_projects.iter_rows(min_row=2, max_row=len(project_data) + 1):
    for cell in row:
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="left", vertical="center")
        # Format budget column as currency
        if cell.column == 2:  # Budget column
            cell.number_format = '$#,##0.00'
            cell.alignment = Alignment(horizontal="right", vertical="center")
        # Format date column
        elif cell.column == 4:  # StartDate column
            cell.number_format = 'YYYY-MM-DD'

# Adjust column widths
ws_projects.column_dimensions['A'].width = 35
ws_projects.column_dimensions['B'].width = 15
ws_projects.column_dimensions['C'].width = 15
ws_projects.column_dimensions['D'].width = 15
ws_projects.column_dimensions['E'].width = 50

# Save the workbook
output_path = "sample_resources.xlsx"
wb.save(output_path)
print(f"✅ Sample Excel file created: {output_path}")
print(f"   - Resources sheet: {len(resource_data)} sample resources")
print(f"   - Projects sheet: {len(project_data)} sample projects")
