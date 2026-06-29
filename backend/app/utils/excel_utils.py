"""Excel file parsing utilities for import functionality"""
import logging
from typing import List, Dict, Any, Tuple, Optional
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelParsingError(Exception):
    """Exception raised during Excel file parsing"""
    pass


class ExcelValidationError(Exception):
    """Exception raised during Excel file validation"""
    pass


class ExcelFormatError(Exception):
    """Exception raised when Excel file format is invalid"""
    pass


class ExcelParser:
    """Utility class for parsing Excel files"""
    
    SUPPORTED_FORMATS = {'.xlsx', '.xls'}
    MAX_PREVIEW_ROWS = 100
    
    @staticmethod
    def validate_file_format(filename: str) -> bool:
        """
        Validate that file has a supported Excel format.
        
        Args:
            filename: The uploaded file name
            
        Returns:
            True if format is supported (.xlsx, .xls)
            
        Raises:
            ExcelParsingError: If format is not supported
        """
        import os
        _, ext = os.path.splitext(filename)
        
        if ext.lower() not in ExcelParser.SUPPORTED_FORMATS:
            supported = ', '.join(ExcelParser.SUPPORTED_FORMATS)
            raise ExcelParsingError(
                f"Unsupported file format '{ext}'. Supported formats: {supported}"
            )
        return True
    
    @staticmethod
    def read_excel_file(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]], int]:
        """
        Read Excel file and extract headers and data rows.
        
        Args:
            file_content: Binary content of Excel file
            
        Returns:
            Tuple of:
            - headers: List of column header names (from first row)
            - rows: List of dictionaries mapping column name to value
            - total_rows: Total number of data rows (excluding header)
            
        Raises:
            ExcelParsingError: If file cannot be parsed
        """
        try:
            # Load workbook from bytes
            file_obj = BytesIO(file_content)
            workbook = openpyxl.load_workbook(file_obj, data_only=True)
            
            # Get active sheet
            sheet = workbook.active
            if sheet is None:
                raise ExcelParsingError("No active worksheet found in Excel file")
            
            # Extract headers from first row
            headers = []
            for cell in sheet[1]:
                value = cell.value
                if value is not None:
                    headers.append(str(value).strip())
                else:
                    headers.append("")
            
            if not headers or all(not h for h in headers):
                raise ExcelParsingError("Excel file has no headers in first row")
            
            # Extract data rows
            rows = []
            total_rows = 0
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=False), start=2):
                # Count total rows
                total_rows = row_idx - 1
                
                # Extract only first 100 rows for preview
                if len(rows) < ExcelParser.MAX_PREVIEW_ROWS:
                    row_data = {}
                    for col_idx, cell in enumerate(row):
                        if col_idx < len(headers):
                            header = headers[col_idx]
                            if header:  # Only include non-empty headers
                                # Get cell value, handling different data types
                                value = cell.value
                                row_data[header] = value
                    
                    # Only add row if it has at least one non-empty cell
                    if any(v is not None for v in row_data.values()):
                        rows.append(row_data)
            
            workbook.close()
            
            return headers, rows, total_rows
            
        except openpyxl.utils.exceptions.InvalidFileException as e:
            raise ExcelParsingError(f"Invalid Excel file format: {str(e)}")
        except Exception as e:
            logger.error(f"Error parsing Excel file: {str(e)}", exc_info=True)
            raise ExcelParsingError(f"Failed to parse Excel file: {str(e)}")
    
    @staticmethod
    def get_column_names_from_headers(headers: List[str]) -> Dict[str, str]:
        """
        Create a mapping of lowercase column names to original headers.
        This helps with case-insensitive column matching.
        
        Args:
            headers: List of column header names
            
        Returns:
            Dictionary mapping lowercase headers to original headers
        """
        mapping = {}
        for header in headers:
            if header:
                mapping[header.lower()] = header
        return mapping
    
    @staticmethod
    def match_column_to_field(
        column_name: str,
        asset_type_fields: List[Dict[str, Any]]
    ) -> Optional[Dict[str, Any]]:
        """
        Match an Excel column name to an asset type field.
        Uses case-insensitive, fuzzy matching.
        
        Args:
            column_name: Name of column from Excel file
            asset_type_fields: List of field definitions from asset type schema
            
        Returns:
            Matching field definition, or None if no match found
        """
        column_lower = column_name.lower().strip()
        
        # Try exact match first (case-insensitive)
        for field in asset_type_fields:
            if field.get('name', '').lower() == column_lower:
                return field
        
        # Try removing underscores/spaces
        column_normalized = column_lower.replace('_', '').replace(' ', '')
        for field in asset_type_fields:
            field_normalized = field.get('name', '').lower().replace('_', '').replace(' ', '')
            if field_normalized == column_normalized:
                return field
        
        return None


def validate_excel_file(file_content: bytes, filename: str) -> Tuple[List[str], List[Dict[str, Any]], int]:
    """
    Validate and parse an Excel file.
    
    Args:
        file_content: Binary content of Excel file
        filename: Name of the uploaded file
        
    Returns:
        Tuple of (headers, rows, total_rows)
        
    Raises:
        ExcelParsingError: If file is invalid or cannot be parsed
    """
    # Validate file format
    ExcelParser.validate_file_format(filename)
    
    # Read and parse file
    return ExcelParser.read_excel_file(file_content)


# Module-level functions for backward compatibility
def read_excel_file(file_content: bytes) -> Tuple[List[str], List[Dict[str, Any]], int]:
    """Read Excel file - module-level function"""
    return ExcelParser.read_excel_file(file_content)


def parse_excel_data(file_content: bytes, filename: str) -> Dict[str, Any]:
    """Parse Excel data from file content"""
    ExcelParser.validate_file_format(filename)
    headers, rows, total_rows = ExcelParser.read_excel_file(file_content)
    return {
        "headers": headers,
        "rows": rows,
        "total_rows": total_rows,
    }


def get_sheet_names(file_content: bytes) -> List[str]:
    """Get list of sheet names from Excel file"""
    try:
        file_obj = BytesIO(file_content)
        workbook = openpyxl.load_workbook(file_obj)
        sheet_names = workbook.sheetnames
        workbook.close()
        return sheet_names
    except Exception as e:
        raise ExcelParsingError(f"Failed to get sheet names: {str(e)}")


def get_sheet_preview(file_content: bytes, sheet_name: str = None, max_rows: int = 10) -> Dict[str, Any]:
    """Get preview of a specific sheet"""
    try:
        file_obj = BytesIO(file_content)
        workbook = openpyxl.load_workbook(file_obj, data_only=True)
        
        if sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.active
        
        # Get headers
        headers = []
        for cell in sheet[1]:
            value = cell.value
            if value is not None:
                headers.append(str(value).strip())
            else:
                headers.append("")
        
        # Get preview rows
        rows = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            if row_idx - 1 >= max_rows:
                break
            row_data = {}
            for col_idx, value in enumerate(row):
                if col_idx < len(headers) and headers[col_idx]:
                    row_data[headers[col_idx]] = value
            if any(v is not None for v in row_data.values()):
                rows.append(row_data)
        
        workbook.close()
        
        return {
            "headers": headers,
            "rows": rows,
            "sheet_name": sheet_name or sheet.title,
        }
    except Exception as e:
        raise ExcelParsingError(f"Failed to get sheet preview: {str(e)}")
