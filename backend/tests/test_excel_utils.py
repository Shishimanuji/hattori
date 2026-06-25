"""Tests for Excel parsing utilities"""

import io
import pytest
from pathlib import Path
from unittest.mock import Mock
import openpyxl
from openpyxl.styles import Font

from app.utils.excel_utils import (
    validate_excel_file,
    read_excel_file,
    parse_excel_data,
    get_sheet_names,
    get_sheet_preview,
    ExcelValidationError,
    ExcelFormatError,
    ExcelParsingError,
)


# ==================== Fixtures ====================

@pytest.fixture
def sample_xlsx_bytes():
    """Create sample XLSX file in memory"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # Add headers
    ws['A1'] = "Name"
    ws['B1'] = "Email"
    ws['C1'] = "Cost"
    
    # Add data rows
    ws['A2'] = "John Doe"
    ws['B2'] = "john@example.com"
    ws['C2'] = 5000
    
    ws['A3'] = "Jane Smith"
    ws['B3'] = "jane@example.com"
    ws['C3'] = 7500
    
    # Add another sheet
    ws2 = wb.create_sheet("Sheet2")
    ws2['A1'] = "Project"
    ws2['B1'] = "Budget"
    ws2['A2'] = "Project A"
    ws2['B2'] = 100000
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_xlsx_no_data_bytes():
    """Create XLSX file with only headers, no data rows"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = "Name"
    ws['B1'] = "Email"
    ws['C1'] = "Cost"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_xlsx_empty_headers_bytes():
    """Create XLSX file with empty first row"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A2'] = "Data"
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


@pytest.fixture
def sample_xlsx_file_path(sample_xlsx_bytes, tmp_path):
    """Create temporary XLSX file"""
    file_path = tmp_path / "test.xlsx"
    file_path.write_bytes(sample_xlsx_bytes)
    return str(file_path)


@pytest.fixture
def mock_upload_file_valid(sample_xlsx_bytes):
    """Create mock UploadFile with valid Excel"""
    mock_file = Mock()
    mock_file.filename = "test.xlsx"
    mock_file.size = len(sample_xlsx_bytes)
    mock_file.file = io.BytesIO(sample_xlsx_bytes)
    return mock_file


@pytest.fixture
def mock_upload_file_invalid_ext():
    """Create mock UploadFile with invalid extension"""
    mock_file = Mock()
    mock_file.filename = "test.txt"
    mock_file.size = 100
    mock_file.file = io.BytesIO(b"invalid")
    return mock_file


@pytest.fixture
def mock_upload_file_too_large():
    """Create mock UploadFile that's too large"""
    mock_file = Mock()
    mock_file.filename = "test.xlsx"
    mock_file.size = 60 * 1024 * 1024 + 1  # Over 50MB limit
    mock_file.file = io.BytesIO(b"x")
    return mock_file


# ==================== validate_excel_file Tests ====================

class TestValidateExcelFile:
    """Tests for validate_excel_file function"""
    
    def test_validate_valid_file(self, mock_upload_file_valid):
        """Test validation of valid Excel file"""
        result = validate_excel_file(mock_upload_file_valid)
        assert result is True
    
    def test_validate_invalid_extension(self, mock_upload_file_invalid_ext):
        """Test validation rejects invalid file extension"""
        with pytest.raises(ExcelValidationError, match="Unsupported file format"):
            validate_excel_file(mock_upload_file_invalid_ext)
    
    def test_validate_file_too_large(self, mock_upload_file_too_large):
        """Test validation rejects files exceeding size limit"""
        with pytest.raises(ExcelValidationError, match="exceeds maximum limit"):
            validate_excel_file(mock_upload_file_too_large)
    
    def test_validate_corrupted_file(self):
        """Test validation rejects corrupted Excel file"""
        mock_file = Mock()
        mock_file.filename = "test.xlsx"
        mock_file.size = 100
        mock_file.file = io.BytesIO(b"not a valid excel file")
        
        with pytest.raises(ExcelFormatError):
            validate_excel_file(mock_file)
    
    def test_validate_empty_file(self):
        """Test validation rejects empty file"""
        mock_file = Mock()
        mock_file.filename = "test.xlsx"
        mock_file.size = 0
        mock_file.file = io.BytesIO(b"")
        
        with pytest.raises(ExcelFormatError):
            validate_excel_file(mock_file)
    
    def test_validate_xls_extension(self):
        """Test validation accepts .xls files"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Header"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        mock_file = Mock()
        mock_file.filename = "test.xls"
        mock_file.size = len(output.getvalue())
        mock_file.file = output
        
        result = validate_excel_file(mock_file)
        assert result is True
    
    def test_validate_xlsm_extension(self):
        """Test validation accepts .xlsm files"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Header"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        mock_file = Mock()
        mock_file.filename = "test.xlsm"
        mock_file.size = len(output.getvalue())
        mock_file.file = output
        
        result = validate_excel_file(mock_file)
        assert result is True


# ==================== read_excel_file Tests ====================

class TestReadExcelFile:
    """Tests for read_excel_file function"""
    
    def test_read_valid_file(self, sample_xlsx_file_path):
        """Test reading valid Excel file"""
        headers, rows = read_excel_file(sample_xlsx_file_path)
        
        assert len(headers) == 3
        assert headers[0] == "Name"
        assert headers[1] == "Email"
        assert headers[2] == "Cost"
        
        assert len(rows) == 2
        assert rows[0][0] == "John Doe"
        assert rows[1][0] == "Jane Smith"
    
    def test_read_specific_sheet(self, sample_xlsx_file_path):
        """Test reading specific sheet by name"""
        headers, rows = read_excel_file(sample_xlsx_file_path, sheet_name="Sheet2")
        
        assert headers[0] == "Project"
        assert headers[1] == "Budget"
        assert rows[0][0] == "Project A"
    
    def test_read_nonexistent_sheet(self, sample_xlsx_file_path):
        """Test reading nonexistent sheet raises error"""
        with pytest.raises(ExcelParsingError, match="Sheet .* not found"):
            read_excel_file(sample_xlsx_file_path, sheet_name="NonExistent")
    
    def test_read_nonexistent_file(self, tmp_path):
        """Test reading nonexistent file raises error"""
        with pytest.raises(ExcelValidationError, match="File not found"):
            read_excel_file(str(tmp_path / "nonexistent.xlsx"))
    
    def test_read_invalid_extension(self, tmp_path):
        """Test reading file with invalid extension raises error"""
        invalid_file = tmp_path / "test.txt"
        invalid_file.write_text("invalid")
        
        with pytest.raises(ExcelValidationError, match="Unsupported file format"):
            read_excel_file(str(invalid_file))
    
    def test_read_file_with_empty_rows(self, tmp_path):
        """Test that empty rows are skipped"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['B1'] = "Value"
        ws['A2'] = "Row1"
        ws['B2'] = 100
        # Row 3 is empty
        ws['A4'] = "Row2"
        ws['B4'] = 200
        
        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        
        headers, rows = read_excel_file(str(file_path))
        assert len(rows) == 2
        assert rows[0][0] == "Row1"
        assert rows[1][0] == "Row2"
    
    def test_read_file_with_null_values(self, tmp_path):
        """Test that null values are converted to empty strings"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['B1'] = "Value"
        ws['A2'] = "Row1"
        ws['B2'] = None
        
        file_path = tmp_path / "test.xlsx"
        wb.save(file_path)
        
        headers, rows = read_excel_file(str(file_path))
        assert rows[0][1] == ""


# ==================== parse_excel_data Tests ====================

class TestParseExcelData:
    """Tests for parse_excel_data function"""
    
    def test_parse_valid_data(self, sample_xlsx_bytes):
        """Test parsing valid Excel data from bytes"""
        headers, rows = parse_excel_data(sample_xlsx_bytes)
        
        assert len(headers) == 3
        assert headers == ["Name", "Email", "Cost"]
        assert len(rows) == 2
        assert rows[0] == ["John Doe", "john@example.com", 5000]
    
    def test_parse_specific_sheet(self, sample_xlsx_bytes):
        """Test parsing specific sheet from bytes"""
        headers, rows = parse_excel_data(sample_xlsx_bytes, sheet_name="Sheet2")
        
        assert headers[0] == "Project"
        assert rows[0][0] == "Project A"
    
    def test_parse_empty_bytes(self):
        """Test parsing empty bytes raises error"""
        with pytest.raises(ExcelValidationError, match="empty"):
            parse_excel_data(b"")
    
    def test_parse_invalid_data(self):
        """Test parsing invalid data raises error"""
        with pytest.raises(ExcelFormatError):
            parse_excel_data(b"not valid excel data")
    
    def test_parse_too_large_data(self):
        """Test parsing data exceeding size limit raises error"""
        large_data = b"x" * (60 * 1024 * 1024)
        with pytest.raises(ExcelValidationError, match="exceeds maximum"):
            parse_excel_data(large_data)
    
    def test_parse_no_data_rows(self, sample_xlsx_no_data_bytes):
        """Test parsing file with only headers"""
        headers, rows = parse_excel_data(sample_xlsx_no_data_bytes)
        
        assert len(headers) == 3
        assert len(rows) == 0
    
    def test_parse_empty_headers_raises_error(self, sample_xlsx_empty_headers_bytes):
        """Test parsing file with empty headers raises error"""
        with pytest.raises(ExcelFormatError, match="no headers"):
            parse_excel_data(sample_xlsx_empty_headers_bytes)
    
    def test_parse_converts_none_to_empty_string(self):
        """Test that None values are converted to empty strings"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Col1"
        ws['B1'] = "Col2"
        ws['A2'] = "Value"
        ws['B2'] = None
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, rows = parse_excel_data(output.getvalue())
        assert rows[0][1] == ""


# ==================== get_sheet_names Tests ====================

class TestGetSheetNames:
    """Tests for get_sheet_names function"""
    
    def test_get_sheet_names(self, sample_xlsx_bytes):
        """Test getting sheet names from Excel file"""
        sheet_names = get_sheet_names(sample_xlsx_bytes)
        
        assert len(sheet_names) >= 2
        assert "Sheet1" in sheet_names or sheet_names[0] == "Sheet1"
    
    def test_get_sheet_names_single_sheet(self):
        """Test getting sheet names from file with single sheet"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Resources"
        ws['A1'] = "Name"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        sheet_names = get_sheet_names(output.getvalue())
        assert sheet_names == ["Resources"]
    
    def test_get_sheet_names_invalid_data(self):
        """Test getting sheet names from invalid data raises error"""
        with pytest.raises(ExcelFormatError):
            get_sheet_names(b"invalid data")


# ==================== get_sheet_preview Tests ====================

class TestGetSheetPreview:
    """Tests for get_sheet_preview function"""
    
    def test_get_preview_valid_data(self, sample_xlsx_bytes):
        """Test getting preview of valid Excel data"""
        headers, preview_rows = get_sheet_preview(sample_xlsx_bytes, max_rows=100)
        
        assert len(headers) == 3
        assert len(preview_rows) <= 100
        assert preview_rows[0][0] == "John Doe"
    
    def test_get_preview_respects_max_rows(self):
        """Test preview respects maximum row limit"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "ID"
        ws['B1'] = "Value"
        
        # Add 200 rows of data
        for i in range(2, 202):
            ws[f'A{i}'] = i - 1
            ws[f'B{i}'] = f"Value_{i-1}"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, preview_rows = get_sheet_preview(output.getvalue(), max_rows=50)
        assert len(preview_rows) == 50
    
    def test_get_preview_specific_sheet(self, sample_xlsx_bytes):
        """Test getting preview of specific sheet"""
        headers, preview_rows = get_sheet_preview(
            sample_xlsx_bytes,
            sheet_name="Sheet2",
            max_rows=100
        )
        
        assert headers[0] == "Project"
    
    def test_get_preview_nonexistent_sheet(self, sample_xlsx_bytes):
        """Test getting preview of nonexistent sheet raises error"""
        with pytest.raises(ExcelFormatError, match="not found"):
            get_sheet_preview(
                sample_xlsx_bytes,
                sheet_name="NonExistent"
            )
    
    def test_get_preview_invalid_data(self):
        """Test getting preview from invalid data raises error"""
        with pytest.raises(ExcelFormatError):
            get_sheet_preview(b"invalid data")
    
    def test_get_preview_skips_empty_rows(self):
        """Test that preview skips empty rows"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Col1"
        ws['A2'] = "Value1"
        # Row 3 is empty
        ws['A4'] = "Value2"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, preview_rows = get_sheet_preview(output.getvalue())
        assert len(preview_rows) == 2
        assert preview_rows[0][0] == "Value1"
        assert preview_rows[1][0] == "Value2"


# ==================== Integration Tests ====================

class TestExcelUtilsIntegration:
    """Integration tests combining multiple functions"""
    
    def test_validate_then_parse_workflow(self, sample_xlsx_bytes):
        """Test typical workflow: validate then parse"""
        mock_file = Mock()
        mock_file.filename = "test.xlsx"
        mock_file.size = len(sample_xlsx_bytes)
        mock_file.file = io.BytesIO(sample_xlsx_bytes)
        
        # Validate first
        assert validate_excel_file(mock_file) is True
        
        # Then parse
        mock_file.file.seek(0)
        file_bytes = mock_file.file.read()
        headers, rows = parse_excel_data(file_bytes)
        
        assert len(headers) == 3
        assert len(rows) == 2
    
    def test_preview_then_import_workflow(self, sample_xlsx_bytes):
        """Test workflow: preview then full import"""
        # Get preview first
        headers, preview_rows = get_sheet_preview(sample_xlsx_bytes, max_rows=10)
        assert len(preview_rows) == 2
        
        # Then get full data
        headers_full, rows_full = parse_excel_data(sample_xlsx_bytes)
        assert len(rows_full) == 2
    
    def test_multi_sheet_workflow(self, sample_xlsx_bytes):
        """Test getting available sheets and parsing each"""
        sheet_names = get_sheet_names(sample_xlsx_bytes)
        assert len(sheet_names) >= 1
        
        for sheet_name in sheet_names:
            headers, rows = parse_excel_data(sample_xlsx_bytes, sheet_name=sheet_name)
            assert len(headers) > 0


# ==================== Edge Cases ====================

class TestEdgeCases:
    """Tests for edge cases and boundary conditions"""
    
    def test_parse_numeric_data_types(self):
        """Test parsing preserves numeric data types"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Integer"
        ws['B1'] = "Float"
        ws['C1'] = "String"
        ws['A2'] = 42
        ws['B2'] = 3.14
        ws['C2'] = "text"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, rows = parse_excel_data(output.getvalue())
        assert rows[0][0] == 42
        assert rows[0][1] == 3.14
        assert rows[0][2] == "text"
    
    def test_parse_boolean_values(self):
        """Test parsing boolean values"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "IsActive"
        ws['A2'] = True
        ws['A3'] = False
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, rows = parse_excel_data(output.getvalue())
        assert rows[0][0] is True
        assert rows[1][0] is False
    
    def test_parse_special_characters_in_headers(self):
        """Test parsing headers with special characters"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Name (Full)"
        ws['B1'] = "Email@Address"
        ws['C1'] = "Cost ($)"
        ws['A2'] = "John"
        ws['B2'] = "john@test.com"
        ws['C2'] = 100
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, rows = parse_excel_data(output.getvalue())
        assert "Name (Full)" in headers[0]
    
    def test_parse_unicode_characters(self):
        """Test parsing unicode characters"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "Name"
        ws['A2'] = "José García"
        ws['A3'] = "中文名字"
        ws['A4'] = "Ελληνικά"
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        headers, rows = parse_excel_data(output.getvalue())
        assert "José García" in rows[0]
        assert "中文名字" in rows[1]
