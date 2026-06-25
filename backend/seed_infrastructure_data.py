"""
Seed infrastructure asset data from Excel workbook
"""
import openpyxl
from datetime import datetime, timedelta
from app.core.database import SessionLocal
from app.models.project import Project
from app.models.resource import Resource
from app.models.asset_type import AssetType, CustomField
from app.models.user import User
import json

# Map Excel columns to database fields
ASSET_CATEGORIES = {
    'Servers': {
        'icon': '🖥️',
        'fields': ['CPU Type', 'CPU Cores', 'RAM (GB)', 'Storage (GB)', 'GPU', 'OS', 'Service Tag', 'Warranty End', 'Antivirus Expiry']
    },
    'Storage': {
        'icon': '💾',
        'fields': ['Storage Type', 'Capacity (TB)', 'OS', 'Service Tag', 'Warranty End', 'Antivirus Expiry']
    },
    'Firewall': {
        'icon': '🔥',
        'fields': ['Model', 'Firmware', 'License Expiry', 'Service Tag', 'Warranty End']
    },
    'L3 Switch': {
        'icon': '🔀',
        'fields': ['Model', 'Port Count', 'Service Tag', 'Warranty End']
    },
    'L2 Switch': {
        'icon': '🔀',
        'fields': ['Model', 'Port Count', 'Service Tag', 'Warranty End']
    },
    'KVM Switch': {
        'icon': '⌨️',
        'fields': ['Model', 'Port Count', 'Service Tag', 'Warranty End']
    },
    'Workstation': {
        'icon': '💻',
        'fields': ['CPU Type', 'CPU Cores', 'RAM (GB)', 'GPU', 'OS', 'Service Tag', 'Warranty End', 'Antivirus Expiry']
    }
}

def parse_date(date_str):
    """Parse date from Excel"""
    if not date_str or date_str == 'N/A':
        return None
    try:
        return datetime.strptime(str(date_str).strip(), '%m/%d/%Y').date()
    except:
        return None

def calculate_health_score(asset_data):
    """Calculate infrastructure health score based on warranty, AV, OS support"""
    score = 100
    
    # Warranty check
    warranty_date = asset_data.get('warranty_date')
    if warranty_date:
        days_until_expiry = (warranty_date - datetime.now().date()).days
        if days_until_expiry < 0:
            score -= 20  # Expired
        elif days_until_expiry < 30:
            score -= 10  # Expiring soon
    
    # Antivirus check
    av_expiry = asset_data.get('av_expiry')
    if av_expiry:
        days_until_expiry = (av_expiry - datetime.now().date()).days
        if days_until_expiry < 0:
            score -= 15  # Expired
        elif days_until_expiry < 30:
            score -= 5   # Expiring soon
    
    return max(0, score)

def seed_data():
    """Load Excel data and seed database"""
    db = SessionLocal()
    
    try:
        # Load workbook
        wb = openpyxl.load_workbook('sample_data/data.xlsx')
        
        # Create projects
        projects_data = {
            'SEA': {'name': 'SEA Infrastructure', 'budget': 500000, 'description': 'Southeast Asia Data Center'},
            'Radar': {'name': 'Radar Infrastructure', 'budget': 300000, 'description': 'Radar Systems Infrastructure'},
            'Security': {'name': 'Security Infrastructure', 'budget': 200000, 'description': 'Security Systems Infrastructure'},
        }
        
        projects = {}
        for project_key, project_info in projects_data.items():
            project = db.query(Project).filter(Project.name == project_info['name']).first()
            if not project:
                project = Project(
                    name=project_info['name'],
                    description=project_info['description'],
                    total_budget=project_info['budget'],
                    status='Active'
                )
                db.add(project)
                db.flush()
            projects[project_key] = project
        
        # Create asset types with custom fields
        for asset_type_name, type_config in ASSET_CATEGORIES.items():
            asset_type = db.query(AssetType).filter(AssetType.name == asset_type_name).first()
            if not asset_type:
                asset_type = AssetType(
                    name=asset_type_name,
                    description=f'{asset_type_name} asset type',
                    is_active=True
                )
                db.add(asset_type)
                db.flush()
                
                # Add custom fields
                field_configs = {
                    'Service Tag': {'type': 'text', 'required': False},
                    'Warranty End': {'type': 'date', 'required': False},
                    'Antivirus Expiry': {'type': 'date', 'required': False},
                    'CPU Cores': {'type': 'number', 'required': False},
                    'RAM (GB)': {'type': 'number', 'required': False},
                    'Storage (GB)': {'type': 'number', 'required': False},
                    'GPU': {'type': 'text', 'required': False},
                    'OS': {'type': 'text', 'required': False},
                    'Room': {'type': 'text', 'required': False},
                    'Custodian': {'type': 'text', 'required': False},
                }
                
                for field_name, field_config in field_configs.items():
                    custom_field = CustomField(
                        asset_type_id=asset_type.id,
                        name=field_name,
                        field_type=field_config['type'],
                        is_required=field_config['required'],
                        order=list(field_configs.keys()).index(field_name)
                    )
                    db.add(custom_field)
        
        db.commit()
        
        # Parse each sheet for assets
        sheet_to_category = {
            'Servers': 'Servers',
            'Storage': 'Storage',
            'Firewall': 'Firewall',
            'L3_Switches': 'L3 Switch',
            'L2_Switches': 'L2 Switch',
            'KVM': 'KVM Switch',
            'Workstations': 'Workstation',
        }
        
        for sheet_name, asset_type_name in sheet_to_category.items():
            if sheet_name not in wb.sheetnames:
                continue
            
            ws = wb[sheet_name]
            asset_type = db.query(AssetType).filter(AssetType.name == asset_type_name).first()
            
            # Get header row
            headers = []
            for cell in ws[1]:
                headers.append(cell.value)
            
            # Process each row
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:  # Skip empty rows
                    continue
                
                asset_name = str(row[0])
                project_key = str(row[1]).strip() if row[1] else 'SEA'
                project = projects.get(project_key, projects['SEA'])
                
                # Extract custom fields
                custom_data = {}
                for col_idx, value in enumerate(row):
                    if col_idx < len(headers) and headers[col_idx]:
                        header = str(headers[col_idx]).strip()
                        if header and value:
                            custom_data[header] = value
                
                # Create resource
                resource = db.query(Resource).filter(
                    Resource.name == asset_name,
                    Resource.asset_type_id == asset_type.id
                ).first()
                
                if not resource:
                    resource = Resource(
                        name=asset_name,
                        asset_type_id=asset_type.id,
                        project_id=project.id,
                        status='Active',
                        cost=0,
                        custom_field_values=json.dumps(custom_data)
                    )
                    db.add(resource)
        
        db.commit()
        print("✅ Infrastructure data seeded successfully!")
        
    except Exception as e:
        print(f"❌ Error seeding data: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == '__main__':
    seed_data()
